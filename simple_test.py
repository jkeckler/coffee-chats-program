import openpyxl

def test_excel_reading():
    filepath = "data/Coffee_Chats_ParticipantsList.xlsx"
    print(f"Testing file: {filepath}")
    
    wb = openpyxl.load_workbook(filepath)
    print(f"Sheets in workbook: {wb.sheetnames}")
    
    if 'Participants' in wb.sheetnames:
        ws = wb['Participants']
        print(f"Participants sheet found. Dimensions: {ws.dimensions}")
        
        # Print headers
        headers = [cell.value for cell in ws[1]]
        print(f"Headers: {headers}")
        
        # Print first few data rows
        for row_idx in range(2, 10):  # Row 2-9
            row_data = []
            for col_idx in range(1, 10):  # First 9 columns
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(f"{cell.value}")
            print(f"Row {row_idx}: {row_data}")
    else:
        print("No 'Participants' sheet found.")

if __name__ == "__main__":
    test_excel_reading()