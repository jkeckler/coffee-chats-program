import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import random
from typing import List, Dict
import os

class ExcelTemplateHandler:
    def __init__(self):
        # Global company structure
        self.locations = {
            "United States": ["UTC-8", "UTC-7", "UTC-5", "UTC-4"],  # West Coast, Mountain, Central, East Coast
            "Canada": ["UTC-7", "UTC-5", "UTC-4"],
            "United Kingdom": ["UTC+0"],
            "Germany": ["UTC+1"],
            "France": ["UTC+1"],
            "Spain": ["UTC+1"],
            "India": ["UTC+5:30"],
            "Singapore": ["UTC+8"],
            "Japan": ["UTC+9"],
            "Australia": ["UTC+10", "UTC+11"],
            "Brazil": ["UTC-3"],
            "Mexico": ["UTC-6"],
            "Netherlands": ["UTC+1"],
            "Sweden": ["UTC+1"],
            "Ireland": ["UTC+0"]
        }

        # Will need to be updated with all the UTCs and the correct departments. 
        
        self.departments = [
            "Engineering",
            "Product Management",
            "Design",
            "Marketing",
            "Sales",
            "Customer Success",
            "Human Resources",
            "Finance",
            "Operations",
            "Data Science",
            "Research"
        ]
        
        self.tenure_options = [
            "Less than 1 year",
            "1 to 3 years",
            "3 to 5 years",
            "5+ years"
        ]

    def generate_sample_data(self, num_employees: int = 104) -> List[Dict]:
        sample_data = []
        
        # Distribution weights
        dept_weights = [25, 15, 10, 10, 10, 8, 5, 5, 5, 4, 3]  # Engineering heavy
        tenure_weights = [30, 40, 20, 10]  # More newer employees
        
        for i in range(num_employees):
            # Generate employee ID with department prefix
            dept = random.choices(self.departments, weights=dept_weights)[0]
            dept_prefix = ''.join(word[0] for word in dept.split())
            emp_id = f"{dept_prefix}{str(i+1).zfill(4)}"
            
            # Select country and timezone
            country = random.choice(list(self.locations.keys()))
            timezone = random.choice(self.locations[country])
            
            # Generate availability based on timezone
            availability = self.generate_realistic_availability(timezone)
            
            employee = {
                "id": emp_id,
                "country": country,
                "department": dept,
                "timezone": timezone,
                "tenure": random.choices(self.tenure_options, weights=tenure_weights)[0],
                "availability": availability
            }
            sample_data.append(employee)
        
        return sample_data

    def generate_realistic_availability(self, timezone: str) -> Dict[int, str]:
        # Convert timezone to rough hour offset
        try:
            offset = int(timezone.replace("UTC", "").split(":")[0])
        except:
            offset = 0
        
        # Generate 9am-5pm availability in local time
        availability = {}
        local_start = 9
        local_end = 17
        
        for hour in range(24):
            # Convert local hour to UTC
            local_hour = (hour - offset) % 24
            # Mark as available during "normal" working hours with some variation
            if local_start <= local_hour <= local_end:
                # 80% chance of being available during core hours
                availability[hour] = 'Y' if random.random() < 0.8 else 'N'
            else:
                # 5% chance of being available outside core hours
                availability[hour] = 'Y' if random.random() < 0.05 else 'N'
        
        return availability

    def create_template(self, output_path: str = "data/participant_template.xlsx") -> None:
        """Create Excel template with sample data."""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Participants"
        
        # Define headers
        basic_headers = [
            "Employee ID",
            "Name",
            "Email",
            "Department",
            "Country",
            "Time Zone",
            "NIQ Tenure",
            "Status"
        ]
        
        # Style definitions
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        time_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add basic info headers
        for col, header in enumerate(basic_headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Add time grid headers
        time_slots = []
        for hour in range(0, 24):
            time_slots.append(f"{hour:02d}:00")
        
        start_col = len(basic_headers) + 1
        for col, time in enumerate(time_slots, start_col):
            cell = ws.cell(row=1, column=col)
            cell.value = time
            cell.font = header_font
            cell.fill = time_fill
            cell.alignment = Alignment(textRotation=90)
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 4
        
        # Generate and add sample data
        sample_data = self.generate_sample_data()
        current_row = 2
        
        for employee in sample_data:
            # Generate name and email
            first_name = f"Employee{current_row-1}"
            last_name = f"{employee['department'].split()[0]}"
            email = f"{first_name.lower()}.{last_name.lower()}@company.com"
            
            row_data = [
                employee['id'],
                f"{first_name} {last_name}",
                email,
                employee['department'],
                employee['country'],
                employee['timezone'],
                employee['tenure'],
                "Active"
            ]
            
            # Add basic info
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
            
            # Add availability
            for hour, available in employee['availability'].items():
                col = start_col + hour
                cell = ws.cell(row=current_row, column=col)
                cell.value = available
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
        
        # Add data validations
        dv_department = DataValidation(type="list", formula1=f'"{",".join(self.departments)}"', allow_blank=False)
        dv_tenure = DataValidation(type="list", formula1=f'"{",".join(self.tenure_options)}"', allow_blank=False)
        dv_status = DataValidation(type="list", formula1='"Active,Opted Out"', allow_blank=False)
        dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
        
        for dv in [dv_department, dv_tenure, dv_status, dv_yn]:
            ws.add_data_validation(dv)
        
        # Apply validations
        dv_department.add(f'D2:D1000')
        dv_tenure.add(f'G2:G1000')
        dv_status.add(f'H2:H1000')
        
        for col in range(start_col, start_col + len(time_slots)):
            dv_yn.add(f'{get_column_letter(col)}2:{get_column_letter(col)}1000')
        
        # Adjust column widths
        for col in range(1, len(basic_headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        wb.save(output_path)
        print(f"Template with sample data created successfully at: {output_path}")

if __name__ == "__main__":
    handler = ExcelTemplateHandler()
    handler.create_template()