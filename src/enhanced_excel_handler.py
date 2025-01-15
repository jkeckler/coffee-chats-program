import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from typing import List, Dict
import os

class ExcelTemplateHandler:
    def create_template(self, output_path: str = "data/participant_template.xlsx") -> None:
        """
        Create a new Excel template file for participant data with time grid availability
        and data validations.
        """
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Participants"
        
        # Define departments
        departments = [
            "Engineering",
            "Product",
            "Design",
            "Marketing",
            "Sales",
            "HR",
            "Finance",
            "Operations"
        ]
        
        # Define tenure options
        tenure_options = [
            "Less than 1 year",
            "1 to 3 years",
            "3 to 5 years",
            "5+ years"
        ]
        
        # Basic info columns
        basic_headers = [
            "Employee ID",
            "Name",
            "Email",
            "Department",
            "Time Zone",
            "NIQ Tenure",
            "Status"
        ]
        
        # Create data validations
        # Department validation
        dv_department = DataValidation(
            type="list",
            formula1=f'"{",".join(departments)}"',
            allow_blank=False
        )
        dv_department.error ='Your entry is not in the list'
        dv_department.errorTitle = 'Invalid Department'
        
        # Tenure validation
        dv_tenure = DataValidation(
            type="list",
            formula1=f'"{",".join(tenure_options)}"',
            allow_blank=False
        )
        dv_tenure.error ='Your entry is not in the list'
        dv_tenure.errorTitle = 'Invalid Tenure'
        
        # Status validation
        dv_status = DataValidation(
            type="list",
            formula1='"Active,Opted Out"',
            allow_blank=False
        )
        
        # Add validations to worksheet
        ws.add_data_validation(dv_department)
        ws.add_data_validation(dv_tenure)
        ws.add_data_validation(dv_status)
        
        # Style definitions
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        time_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add basic info headers
        for col, header in enumerate(basic_headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Add time grid headers
        time_slots = []
        for hour in range(0, 24):
            time_slots.append(f"{hour:02d}:00")
        
        start_col = len(basic_headers) + 1
        for col, time in enumerate(time_slots, start_col):
            cell = ws.cell(row=1, column=col)
            cell.value = time
            cell.font = header_font
            cell.fill = time_fill
            cell.alignment = Alignment(textRotation=90)
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 4
        
        # Add example row
        example_data = [
            "EMP001",
            "John Doe",
            "john.doe@company.com",
            "Engineering",
            "UTC-5",
            "1 to 3 years",
            "Active"
        ]
        
        # Add basic info example
        for col, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = value
            cell.border = border
        
        # Apply data validations to columns
        dv_department.add(f'D2:D1000')  # Department column
        dv_tenure.add(f'F2:F1000')      # Tenure column
        dv_status.add(f'G2:G1000')      # Status column
        
        # Add availability example (Y/N)
        dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
        ws.add_data_validation(dv_yn)
        
        for col in range(start_col, start_col + len(time_slots)):
            cell = ws.cell(row=2, column=col)
            # Example: Available 9:00-17:00
            hour = col - start_col
            cell.value = 'Y' if 9 <= hour <= 17 else 'N'
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            dv_yn.add(get_column_letter(col) + '2:' + get_column_letter(col) + '1000')
        
        # Adjust basic info column widths
        for col in range(1, len(basic_headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Add instruction row
        instruction_row = ws.row_dimensions[3]
        cell = ws.cell(row=3, column=1)
        cell.value = "Instructions: Mark Y for available hours, N for unavailable hours. Use dropdowns for Department, NIQ Tenure, and Status."
        cell.font = Font(italic=True)
        
        wb.save(output_path)
        print(f"Enhanced template created successfully at: {output_path}")

if __name__ == "__main__":
    # Create template
    handler = ExcelTemplateHandler()
    handler.create_template()