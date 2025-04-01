from typing import List, Dict
import openpyxl 
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import os
from src.models import CoffeeGroup
from src.utils import convert_local_to_utc, generate_time_ranges



class ExcelTemplateHandler:
    basic_headers = ["Employee ID", "Name", "Email", "Department", "Country", "Timezone", "Start Date", "Status"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0046BE", end_color="0046BE", fill_type="solid")
    time_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def __init__(self):
        # Single department for your use case
        self.departments = ["Business Intelligence"]
        
        # Updated locations with proper timezone format
        self.locations = {
            "United States": ["UTC-12:00", "UTC-11:00", "UTC-10:00", "UTC-09:00", 
                            "UTC-08:00", "UTC-07:00", "UTC-06:00", "UTC-05:00", "UTC-04:00"],
            "Canada": ["UTC-08:00", "UTC-07:00", "UTC-06:00", "UTC-05:00", "UTC-04:00"],
            "India": ["UTC+05:30"],
            "Colombia": ["UTC-05:00"],
            "United Kingdom": ["UTC+00:00"],
            "France": ["UTC+01:00"],
            "Afghanistan": ["UTC+04:30"],
            "Australia": ["UTC+08:00", "UTC+09:30", "UTC+10:00"],
            "Bangladesh": ["UTC+06:00"],
            "Bhutan": ["UTC+06:00"],
            "Brunei": ["UTC+08:00"],
            "Cambodia": ["UTC+07:00"],
            "China": ["UTC+08:00"],
            "Germany": ["UTC+01:00", "UTC+02:00"],
            "Japan": ["UTC+09:00"],
            "Brazil": ["UTC-05:00", "UTC-04:00", "UTC-03:00", "UTC-02:00"],
            "South Africa": ["UTC+02:00"],
            "Singapore": ["UTC+08:00"],
            "Netherlands": ["UTC+01:00", "UTC+02:00"],
            "Italy": ["UTC+01:00", "UTC+02:00"],
            "Mexico": ["UTC-08:00", "UTC-07:00", "UTC-06:00"]
        }

        # Generate list of unique timezones
        self.timezones = sorted(list(set(tz for tzs in self.locations.values() for tz in tzs)))

        # Basic headers for the Excel template
        self.basic_headers = ["Employee ID", "Name", "Email", "Department", 
                            "Country", "Timezone", "Start Date", "Status"]
        
        # Define styles for Excel formatting
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="0046BE", end_color="0046BE", fill_type="solid")
        self.time_fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    def create_template(self, output_path: str = "data/participant_template.xlsx"):
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb = openpyxl.Workbook()
    
        # Remove the default sheet created by openpyxl
        wb.remove(wb.active)
        
        # Create Instructions sheet
        instructions_ws = wb.create_sheet("Instructions")
        wb.active = instructions_ws  # Make Instructions the active sheet
        self.create_instructions_sheet(instructions_ws)
        
        # Create Participants sheet
        participants_ws = wb.create_sheet("Participants")
        self.create_participants_sheet(participants_ws)

        wb.save(output_path)
        print(f"Template created at: {output_path}")
    def create_instructions_sheet(self, ws):
        # Add logo
        try:
            logo_path = 'assets/nielseniq_logo.png'
            print(f"Attempting to load logo from: {logo_path}")
            img = openpyxl.drawing.image.Image(logo_path)
            img.anchor = 'A1'
            ws.add_image(img)
        except Exception as e:
            print(f"Logo file error: {str(e)}")

        # Add title
        ws['B1'] = "Global Coffee Chats Program - Business Intelligence"
        ws['B1'].font = Font(size=20, bold=True)

        ws['B3'] = "Instructions:"
        ws['B3'].font = Font(size=16, bold=True)
        ws['B5'] = "1. Fill out your personal information in the Participants sheet."
        ws['B6'] = "2. Select your availability for each hour (Y for available, N for unavailable)."
        ws['B7'] = "3. Your total available hours will be calculated automatically."
        ws['B8'] = "4. For support, contact: GlobalBI.Engagement@smb.nielseniq.com"

        # Adjust column width for better readability
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 75
    def create_participants_sheet(self, ws):
        # Add headers
        headers = self.basic_headers + ["Total Available Hours"] + [f"{h:02d}:00" for h in range(24)]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column widths
        for col in range(1, len(self.basic_headers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20
        for col in range(len(self.basic_headers) + 2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 5

        # Add data validation
        self.add_data_validation(ws)

        # Add conditional formatting
        self.add_conditional_formatting(ws)

        # Add total available hours formula
        total_hours_col = len(self.basic_headers) + 1
        availability_start_col = total_hours_col + 1
        for row in range(2, 1002):  # Assuming up to 1000 participants
            cell = ws.cell(row=row, column=total_hours_col)
            cell.value = f'=COUNTIF({get_column_letter(availability_start_col)}{row}:{get_column_letter(availability_start_col+23)}{row},"Y")'

        # Freeze panes
        ws.freeze_panes = 'D2'

        # Protect sheet
        ws.protection.sheet = True
        ws.protection.password = 'coffee_chats'  # Change this to a secure password

        for row in ws['A2:AG1001']:
            for cell in row:
                cell.protection = Protection(locked=False)
    def add_data_validation(self, ws):
        dv_dept = DataValidation(type="list", formula1=f'"{",".join(self.departments)}"', allow_blank=False)
        dv_country = DataValidation(type="list", formula1=f'"{",".join(self.locations.keys())}"', allow_blank=False)
        dv_tz = DataValidation(type="list", formula1=f'"{",".join(self.timezones)}"', allow_blank=False)
        dv_date = DataValidation(type="date", allow_blank=True)
        dv_status = DataValidation(type="list", formula1='"Active,Opted Out"', allow_blank=False)
        dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)

        for dv in [dv_dept, dv_country, dv_tz, dv_date, dv_status, dv_yn]:
            ws.add_data_validation(dv)

        dv_dept.add('D2:D1001')
        dv_country.add('E2:E1001')
        dv_tz.add('F2:F1001')
        dv_date.add('G2:G1001')  # This is now the Start Date column
        dv_status.add('H2:H1001')
        for col in range(9, 33):  # Columns I to AF
            dv_yn.add(f'{get_column_letter(col)}2:{get_column_letter(col)}1001')
    def add_conditional_formatting(self, ws):
        green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        rule = CellIsRule(operator='equal', formula=['"Y"'], fill=green_fill)
        
        for col in range(9, 33):  # Columns I to AF
            ws.conditional_formatting.add(f'{get_column_letter(col)}2:{get_column_letter(col)}1001', rule)
    def validate_employee_id(self, id_value) -> str:
        """
        Validate and format employee ID
        Returns formatted ID or raises ValueError if invalid
        """
        # Debug output to see what's coming in
        print(f"Validating ID: {id_value}, type: {type(id_value)}")
        
        # Handle None or empty values
        if id_value is None:
            raise ValueError("Employee ID cannot be empty")
            
        # Convert to string, removing any spaces
        id_str = str(id_value).strip()
        
        # Handle empty strings
        if not id_str:
            raise ValueError("Employee ID cannot be empty")
        
        # Special case for your exact format which starts with zeros
        if isinstance(id_value, str) and id_value.isdigit():
            return id_value
        
        try:
            # Check if it's a valid number
            id_num = int(id_str)
            
            # Format to 8 digits with leading zeros
            return str(id_num).zfill(8)
            
        except ValueError:
            raise ValueError(f"Invalid Employee ID format: {id_value}")   
    def read_participants_from_excel(self, filepath: str) -> List[Dict]:
        """Read participant data from Excel template and convert to internal format"""
        try:
            print(f"Opening file: {filepath}")
            wb = openpyxl.load_workbook(filepath)
            
            # Print all sheet names
            print(f"Sheet names: {wb.sheetnames}")
            
            ws = wb['Participants']
            participants = []
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            print(f"Found headers: {headers}")
            
            # Print first few rows for debugging
            print("First few rows:")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=5)):
                print(f"Row {row_idx+2}: {[cell.value for cell in row]}")

            print("First few rows with cell details:")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=5)):
                cell_details = []
                for cell in row:
                    cell_details.append(f"Col {cell.column_letter}: {cell.value} (type: {type(cell.value)})")
                print(f"Row {row_idx+2} details: {cell_details}")
            
            for row in ws.iter_rows(min_row=2):
                # Skip rows without an Employee ID
                if row[0].value is None:
                    continue
                    
                # Skip rows that are completely empty
                if not any(cell.value for cell in row):
                    continue
                    
                try:
                    # Validate employee ID first
                    employee_id = self.validate_employee_id(row[0].value)
                    
                    # If ID is valid, create participants
                    participant = {
                        'id': employee_id,
                        'name': row[1].value,
                        'email': row[2].value,
                        'department': row[3].value,
                        'country': row[4].value,
                        'timezone': row[5].value,
                        'availability': {},
                        'captain_count': 0
                    }
                    
                    # Basic validation of required fields
                    if not all([participant['name'], participant['email'], 
                            participant['department'], participant['country'], 
                            participant['timezone']]):
                        print(f"Skipping row with missing required fields: ID {employee_id}")
                        continue
                    
                    # Validate timezone format
                    if not participant['timezone'].startswith('UTC'):
                        print(f"Invalid timezone format for {participant['name']}: {participant['timezone']}")
                        continue
                    
                    # Convert availability to Y/N format
                    availability_start = 9  # Column where availability starts
                    for hour in range(24):
                        col_index = availability_start + hour
                        value = row[col_index].value
                        participant['availability'][str(hour)] = 'Y' if value == 'Y' else 'N'
                    
                    participants.append(participant)
                    
                except ValueError as e:
                    print(f"Skipping row due to validation error: {e}")
                    continue
            
            print(f"Successfully loaded {len(participants)} participants")
            return participants
            
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return []

        def export_matches_to_excel(self, groups: List[CoffeeGroup], output_path: str = "data/coffee_chat_matches.xlsx"):
            """Export matches to Excel with UTC availability ranges"""
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Coffee Chat Groups"

            header_font = Font(bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            headers = ["Group", "Captain", "Captain Dept", "Members", 
                    "Member Departments", "Suggested Meeting Time", "UTC Working Hours"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.border = border

            for idx, group in enumerate(groups, 1):
                if not group.captain:
                    continue
                    
                row = idx + 1
                
                ws.cell(row=row, column=1, value=f"Group {idx}")
                ws.cell(row=row, column=2, value=f"{group.captain['name']} ({group.captain['timezone']})")
                ws.cell(row=row, column=3, value=f"{group.captain['department']}")
                
                other_members = [f"{m['name']} ({m['timezone']})" for m in group.members if m != group.captain]
                ws.cell(row=row, column=4, value="\n".join(other_members))
                
                member_depts = [f"{m['name']}: {m['department']}" for m in group.members if m != group.captain]
                ws.cell(row=row, column=5, value="\n".join(member_depts))
                
                if group.optimal_meeting_time is not None:
                    ws.cell(row=row, column=6, value=f"{group.optimal_meeting_time:02d}:00 UTC")
                
                utc_hours = []
                for member in group.members:
                    member_utc_hours = set()
                    for hour, status in member['availability'].items():
                        if status == 'Y':
                            utc_hour = convert_local_to_utc(hour, member['timezone'])
                            if utc_hour is not None:
                                member_utc_hours.add(utc_hour)
                    time_ranges = generate_time_ranges(sorted(member_utc_hours))
                    utc_hours.append(f"{member['name']}: {time_ranges}")
                    
                ws.cell(row=row, column=7, value="\n".join(utc_hours))

                for col in range(1, 8):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(wrapText=True, vertical='center')

            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    if cell.value:
                        lines = str(cell.value).count('\n') + 1
                        max_length = max(max_length, 
                                    max(len(line) for line in str(cell.value).split('\n')))
                        ws.row_dimensions[cell.row].height = max(15 * lines, 20)
                ws.column_dimensions[column[0].column_letter].width = max_length + 2

            wb.save(output_path)
            print(f"Match results exported to: {output_path}")

def quick_read_participants(filepath: str) -> List[Dict]:
    """Convenience function to read participants without creating a handler manually"""
    handler = ExcelTemplateHandler()
    return handler.read_participants_from_excel(filepath)

def quick_export_matches(groups: List[CoffeeGroup], output_path: str):
    """Quick implementation to export matches to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coffee Chat Groups"
    
    # Add headers
    headers = ["Group", "Captain", "Captain Dept", "Members", 
              "Member Departments", "Suggested Meeting Time", "UTC Working Hours"]
    
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.border = border
    
    # Add data
    for idx, group in enumerate(groups, 1):
        if not group.captain:
            continue
            
        row = idx + 1
        
        ws.cell(row=row, column=1, value=f"Group {idx}")
        ws.cell(row=row, column=2, value=f"{group.captain['name']} ({group.captain['timezone']})")
        ws.cell(row=row, column=3, value=f"{group.captain['department']}")
        
        other_members = [f"{m['name']} ({m['timezone']})" for m in group.members if m != group.captain]
        ws.cell(row=row, column=4, value="\n".join(other_members))
        
        member_depts = [f"{m['name']}: {m['department']}" for m in group.members if m != group.captain]
        ws.cell(row=row, column=5, value="\n".join(member_depts))
        
        if group.optimal_meeting_time is not None:
            ws.cell(row=row, column=6, value=f"{group.optimal_meeting_time:02d}:00 UTC")
        
        utc_hours = []
        for member in group.members:
            member_utc_hours = set()
            for hour, status in member['availability'].items():
                if status == 'Y':
                    utc_hour = convert_local_to_utc(hour, member['timezone'])
                    if utc_hour is not None:
                        member_utc_hours.add(utc_hour)
            time_ranges = generate_time_ranges(sorted(member_utc_hours))
            utc_hours.append(f"{member['name']}: {time_ranges}")
            
        ws.cell(row=row, column=7, value="\n".join(utc_hours))

        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(wrapText=True, vertical='center')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                max_length = max(max_length, 
                              max(len(line) for line in str(cell.value).split('\n')))
                ws.row_dimensions[cell.row].height = max(15 * lines, 20)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2
    
    # Save the file
    wb.save(output_path)
    print(f"Match results exported to: {output_path}")