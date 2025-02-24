from src.excel_handler import ExcelTemplateHandler
import openpyxl
import os

def create_test_excel():
    """Create a test Excel file with various test cases"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants"
    
    # Add headers
    headers = ["Employee ID", "Name", "Email", "Department", "Country", 
              "Timezone", "Start Date", "Status", "Total Available Hours"]
    # Add time slots
    headers.extend([f"{i:02d}:00" for i in range(24)])
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Test cases
    test_data = [
        # Valid entry
        ["00123456", "John Doe", "john@nielseniq.com", "Business Intelligence", 
         "United States", "UTC-05:00", "2024-01-01", "Active", "8"] + ["Y"]*8 + ["N"]*16,
        
        # Invalid ID (too long)
        ["123456789", "Jane Smith", "jane@nielseniq.com", "Business Intelligence",
         "India", "UTC+05:30", "2024-01-01", "Active", "8"] + ["Y"]*8 + ["N"]*16,
        
        # Invalid ID (non-numeric)
        ["ABC123", "Bob Wilson", "bob@nielseniq.com", "Business Intelligence",
         "United Kingdom", "UTC+00:00", "2024-01-01", "Active", "8"] + ["Y"]*8 + ["N"]*16,
        
        # Missing required fields
        ["00123457", "", "", "Business Intelligence",
         "United States", "UTC-05:00", "2024-01-01", "Active", "8"] + ["Y"]*8 + ["N"]*16,
    ]
    
    for row_idx, data in enumerate(test_data, 2):
        for col_idx, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the test file
    test_file_path = "data/test_participants.xlsx"
    os.makedirs("data", exist_ok=True)
    wb.save(test_file_path)
    return test_file_path

def test_participant_reading():
    """Test reading participants with various test cases"""
    handler = ExcelTemplateHandler()
    
    # Create test file
    test_file = create_test_excel()
    print("\nTesting participant reading with various test cases...")
    
    # Read participants
    participants = handler.read_participants_from_excel(test_file)
    
    # Print results
    print(f"\nSuccessfully read {len(participants)} valid participants")
    print("\nValid participant details:")
    for participant in participants:
        print(f"\nID: {participant['id']}")
        print(f"Name: {participant['name']}")
        print(f"Timezone: {participant['timezone']}")
        print(f"Available hours: {sum(1 for v in participant['availability'].values() if v == 'Y')}")

if __name__ == "__main__":
    test_participant_reading()