from src.excel_handler import ExcelTemplateHandler
from src.models import CoffeeGroup
from src.utils import convert_timezone_to_float
import openpyxl
import os

def create_timezone_test_data():
    """Create test Excel file with participants in different time zones"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants"
    
    # Add headers
    headers = ["Employee ID", "Name", "Email", "Department", "Country", 
              "Timezone", "Start Date", "Status", "Total Available Hours"]
    headers.extend([f"{i:02d}:00" for i in range(24)])
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Test participants in different time zones with overlapping times
    test_data = [
        # US Eastern (9am-5pm local = 14:00-22:00 UTC)
        ["00100001", "US East Worker", "east@nielseniq.com", "Business Intelligence",
         "United States", "UTC-05:00", "2024-01-01", "Active", "8"] + 
         ["N"]*9 + ["Y"]*8 + ["N"]*7,
        
        # India (10am-6pm local = 4:30-12:30 UTC)
        ["00100002", "India Worker", "india@nielseniq.com", "Business Intelligence",
         "India", "UTC+05:30", "2024-01-01", "Active", "8"] + 
         ["N"]*4 + ["Y"]*8 + ["N"]*12,
        
        # UK (10am-6pm local = 10:00-18:00 UTC)
        ["00100003", "UK Worker", "uk@nielseniq.com", "Business Intelligence",
         "United Kingdom", "UTC+00:00", "2024-01-01", "Active", "8"] + 
         ["N"]*10 + ["Y"]*8 + ["N"]*6,
    ]
    
    for row_idx, data in enumerate(test_data, 2):
        for col_idx, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    test_file_path = "data/timezone_test.xlsx"
    os.makedirs("data", exist_ok=True)
    wb.save(test_file_path)
    return test_file_path

def test_timezone_matching():
    """Test timezone conversion and matching"""
    handler = ExcelTemplateHandler()
    test_file = create_timezone_test_data()
    
    print("\nTesting timezone conversion and matching...")
    
    # Read participants
    participants = handler.read_participants_from_excel(test_file)
    
    # Create test groups
    print("\nTesting possible group combinations:")
    
    # Test all possible pairs
    for i in range(len(participants)):
        for j in range(i + 1, len(participants)):
            p1 = participants[i]
            p2 = participants[j]
            group = CoffeeGroup([p1, p2])
            
            print(f"\nTesting match: {p1['name']} ({p1['timezone']}) with {p2['name']} ({p2['timezone']})")
            print(f"Valid group: {group.is_valid_group()}")
            if group.is_valid_group():
                print(f"Common hours (UTC): {sorted(group.common_hours)}")
                print(f"Optimal meeting time (UTC): {group.optimal_meeting_time}:00")
    
    # Test a group of three
    print("\nTesting three-person group:")
    group = CoffeeGroup(participants)
    print(f"Valid group: {group.is_valid_group()}")
    if group.is_valid_group():
        print(f"Common hours (UTC): {sorted(group.common_hours)}")
        print(f"Optimal meeting time (UTC): {group.optimal_meeting_time}:00")

if __name__ == "__main__":
    test_timezone_matching()